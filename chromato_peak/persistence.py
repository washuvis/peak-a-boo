"""Durable Excel-backed workflow persistence for the public synthetic demo.

The synthetic HDF5 signal remains immutable. Analyst state is written to the
local working ``synthetic_reference.xlsx`` workbook:

* appended workflow columns in ``reference_peaks`` for linked reference intervals; and
* a ``ReviewActions`` sheet for every queue case, including unmatched detector
  candidates that cannot be represented by an existing reference row.
* a ``ReviewAuditTrail`` append-only sheet for disposition events.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
import os
import shutil
import tempfile
from typing import Any, Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WORKFLOW_LABELS: dict[str, str] = {
    "open": "Needs review",
    "accepted": "Accepted / reviewed",
    "exception": "Exception flagged · active",
    "exception_resolved": "Exception resolved",
}
RESOLVED_STATES = {"accepted", "exception_resolved"}

PRIMARY_WORKFLOW_COLUMNS = [
    "ReviewStatus",
    "ReviewDecision",
    "ReviewResolved",
    "ReviewCaseKey",
    "ReviewRegionId",
    "ReviewTitle",
    "ReviewNote",
    "ReviewUpdatedAt",
]
ACTION_COLUMNS = [
    "CaseKey",
    "ChannelId",
    "RegionId",
    "PeakId",
    "ReferenceExcelRow",
    "TimeStart",
    "TimeEnd",
    "ApexTime",
    "Title",
    "MatchStatus",
    "ActionState",
    "ReviewStatus",
    "ReviewResolved",
    "AnalystNote",
    "UpdatedAt",
]
AUDIT_COLUMNS = ["Timestamp", "ChannelId", "CaseKey", "RegionId", "Action", "Detail"]
ACTION_SHEET = "ReviewActions"
AUDIT_SHEET = "ReviewAuditTrail"


def ensure_working_label_workbook(source_path: str | Path, data_dir: str | Path) -> Path:
    """Return a writable local workbook, copying the supplied source once.

    A workbook supplied directly inside ``data/synthetic_reference.xlsx`` is used in place.
    When the app finds the source elsewhere (for example ``/mnt/data``), it
    creates ``data/synthetic_reference.xlsx`` as a working copy before any analyst writes.
    """
    source = Path(source_path).expanduser().resolve()
    target_dir = Path(data_dir).expanduser().resolve()
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "synthetic_reference.xlsx"
    if target.exists():
        return target
    if source.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        shutil.copy2(source, target)
    elif source.suffix.lower() == ".csv":
        pd.read_csv(source).to_excel(target, sheet_name="reference_peaks", index=False)
    else:
        raise ValueError(f"Unsupported label source: {source.suffix}")
    return target


def _primary_sheet(workbook: Workbook) -> Worksheet:
    return workbook["reference_peaks"] if "reference_peaks" in workbook.sheetnames else workbook[workbook.sheetnames[0]]


def _headers(sheet: Worksheet) -> dict[str, int]:
    return {str(cell.value): int(cell.column) for cell in sheet[1] if cell.value is not None}


def _ensure_columns(sheet: Worksheet, columns: list[str]) -> dict[str, int]:
    mapping = _headers(sheet)
    for name in columns:
        if name not in mapping:
            col = sheet.max_column + 1
            sheet.cell(row=1, column=col, value=name)
            mapping[name] = col
    return mapping


def _ensure_log_sheet(workbook: Workbook, name: str, columns: list[str]) -> Worksheet:
    if name not in workbook.sheetnames:
        ws = workbook.create_sheet(name)
        ws.append(columns)
        return ws
    ws = workbook[name]
    _ensure_columns(ws, columns)
    return ws


def _atomic_save(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(prefix=f".{path.stem}_", suffix=path.suffix, dir=path.parent, delete=False)
    temp_path = Path(handle.name)
    handle.close()
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _finite_int(value: Any) -> int | None:
    try:
        number = float(value)
        return int(number) if np.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def workflow_status(state: str, note: str = "") -> str:
    if state not in WORKFLOW_LABELS:
        raise ValueError(f"Unsupported workflow state: {state}")
    return "Note added · needs review" if state == "open" and note.strip() else WORKFLOW_LABELS[state]


def persist_review_record(
    workbook_path: str | Path,
    case: Mapping[str, Any] | pd.Series,
    *,
    state: str,
    note: str = "",
    action: str,
    detail: str,
) -> dict[str, Any]:
    """Upsert a case disposition, update its linked label row, and append audit history."""
    if state not in WORKFLOW_LABELS:
        raise ValueError(f"Unsupported workflow state: {state}")
    path = Path(workbook_path)
    timestamp = datetime.now().isoformat(timespec="seconds")
    status = workflow_status(state, note)
    resolved = state in RESOLVED_STATES
    case_key = str(case["case_key"])
    channel_id = _finite_int(case.get("channel_id"))
    source_row = _finite_int(case.get("label_source_row"))
    row_data: dict[str, Any] = {
        "CaseKey": case_key,
        "ChannelId": channel_id,
        "RegionId": str(case.get("region_id", "")),
        "PeakId": _finite_int(case.get("peak_id")),
        "ReferenceExcelRow": source_row,
        "TimeStart": float(case.get("time_start", np.nan)),
        "TimeEnd": float(case.get("time_end", np.nan)),
        "ApexTime": float(case.get("time", np.nan)),
        "Title": str(case.get("title", "")),
        "MatchStatus": str(case.get("match_status", "")),
        "ActionState": state,
        "ReviewStatus": status,
        "ReviewResolved": resolved,
        "AnalystNote": str(note).strip(),
        "UpdatedAt": timestamp,
    }
    workbook = load_workbook(path)
    primary = _primary_sheet(workbook)
    primary_cols = _ensure_columns(primary, PRIMARY_WORKFLOW_COLUMNS)
    if source_row is not None and 2 <= source_row <= primary.max_row:
        linked_values = {
            "ReviewStatus": status,
            "ReviewDecision": action,
            "ReviewResolved": resolved,
            "ReviewCaseKey": case_key,
            "ReviewRegionId": row_data["RegionId"],
            "ReviewTitle": row_data["Title"],
            "ReviewNote": row_data["AnalystNote"],
            "ReviewUpdatedAt": timestamp,
        }
        for field, value in linked_values.items():
            primary.cell(row=source_row, column=primary_cols[field], value=value)

    action_sheet = _ensure_log_sheet(workbook, ACTION_SHEET, ACTION_COLUMNS)
    action_headers = _headers(action_sheet)
    target_row = None
    for excel_row in range(2, action_sheet.max_row + 1):
        if str(action_sheet.cell(excel_row, action_headers["CaseKey"]).value) == case_key:
            target_row = excel_row
            break
    if target_row is None:
        target_row = action_sheet.max_row + 1
    for field in ACTION_COLUMNS:
        action_sheet.cell(row=target_row, column=action_headers[field], value=row_data[field])

    audit_sheet = _ensure_log_sheet(workbook, AUDIT_SHEET, AUDIT_COLUMNS)
    audit_headers = _headers(audit_sheet)
    audit_row = audit_sheet.max_row + 1
    audit_values = {
        "Timestamp": timestamp,
        "ChannelId": channel_id,
        "CaseKey": case_key,
        "RegionId": row_data["RegionId"],
        "Action": action,
        "Detail": detail,
    }
    for field in AUDIT_COLUMNS:
        audit_sheet.cell(row=audit_row, column=audit_headers[field], value=audit_values[field])
    _atomic_save(workbook, path)
    return {"state": state, "note": row_data["AnalystNote"], "updated_at": timestamp, **row_data}


def load_review_records(workbook_path: str | Path, channel_id: int | None = None) -> dict[str, dict[str, Any]]:
    """Load current case dispositions from the durable ReviewActions sheet."""
    path = Path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    if ACTION_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[ACTION_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    columns = [str(value) for value in rows[0]]
    result: dict[str, dict[str, Any]] = {}
    for values in rows[1:]:
        record = dict(zip(columns, values))
        if channel_id is not None and _finite_int(record.get("ChannelId")) != int(channel_id):
            continue
        key = str(record.get("CaseKey", "")).strip()
        if not key:
            continue
        result[key] = {
            "state": str(record.get("ActionState") or "open"),
            "note": str(record.get("AnalystNote") or ""),
            "updated_at": str(record.get("UpdatedAt") or ""),
        }
    return result


def load_persisted_audit(workbook_path: str | Path, channel_id: int | None = None, limit: int = 100) -> list[dict[str, str]]:
    """Read recorded disposition history newest-first from the workbook."""
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    if AUDIT_SHEET not in workbook.sheetnames:
        return []
    rows = list(workbook[AUDIT_SHEET].iter_rows(values_only=True))
    if not rows:
        return []
    columns = [str(value) for value in rows[0]]
    log: list[dict[str, str]] = []
    for values in rows[1:]:
        record = dict(zip(columns, values))
        if channel_id is not None and _finite_int(record.get("ChannelId")) != int(channel_id):
            continue
        timestamp = str(record.get("Timestamp") or "")
        log.append({
            "time": timestamp[11:16] if len(timestamp) >= 16 else timestamp,
            "action": str(record.get("Action") or ""),
            "detail": str(record.get("Detail") or ""),
        })
    return list(reversed(log))[:limit]
