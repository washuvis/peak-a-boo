from __future__ import annotations

import json
from pathlib import Path
import shutil

import joblib
import numpy as np
from openpyxl import load_workbook

from chromato_peak.data_io import data_manifest, labels_for_channel, load_label_table
from chromato_peak.persistence import load_persisted_audit, load_review_records, persist_review_record
from chromato_peak.pipeline import run_pipeline
from chromato_peak.review import apply_review_state, build_review_queue, report_json, review_queue_counts
from chromato_peak.visualization import chromatogram_figure, selected_focus_ranges

ROOT = Path(__file__).resolve().parents[1]
H5 = ROOT / "data" / "synthetic_chromatograms.h5"
REF = ROOT / "data" / "synthetic_reference.xlsx"
MODEL = ROOT / "models" / "synthetic_peak_classifier.joblib"


def test_public_bundle_has_only_synthetic_runtime_data():
    assert H5.exists()
    assert REF.exists()
    assert MODEL.exists()


def test_synthetic_manifest_and_reference_schema():
    manifest = data_manifest(H5, REF)
    assert manifest["n_signal_channels"] == 6
    assert manifest["n_reference_rows"] == 84
    assert manifest["common_channels"] == [1001, 1002, 1003, 1004, 1005, 1006]
    labels = load_label_table(REF)
    assert len(labels_for_channel(labels, 1001)) == 14
    assert {"PeakId", "ChannelId", "StartTime", "EndTime", "RetentionTime"}.issubset(labels.columns)


def test_pipeline_and_review_queue_are_nonempty():
    result = run_pipeline(H5, REF, 1001)
    assert len(result.time) == 3201
    assert len(result.final_peaks) > 0
    assert len(result.label_table) == 14
    assert result.final_metrics["TP"] > 0
    queue = build_review_queue(result, max_items=10)
    assert not queue.empty
    assert queue["region_id"].str.match(r"(REF|CAND)-\d+").all()
    assert queue["data_origin"].str.contains("synthetic", case=False).all()


def test_focus_changes_viewport_not_signal_samples():
    result = run_pipeline(H5, REF, 1001)
    queue = apply_review_state(build_review_queue(result, max_items=10), {})
    selected = queue.iloc[0]
    overview_x, overview_y = selected_focus_ranges(result, selected, False)
    focus_x, focus_y = selected_focus_ranges(result, selected, True)
    assert overview_x is None
    assert focus_x is not None
    assert focus_y == overview_y
    overview = chromatogram_figure(
        result, queue, selected,
        show_raw=True, show_band=True, show_labels=True,
        show_official_peaks=True, zoom_selected=False,
        show_review_regions=True,
    )
    focus = chromatogram_figure(
        result, queue, selected,
        show_raw=True, show_band=True, show_labels=True,
        show_official_peaks=True, zoom_selected=True,
        show_review_regions=True,
    )
    raw_overview = next(trace for trace in overview.data if trace.name == "Raw signal")
    raw_focus = next(trace for trace in focus.data if trace.name == "Raw signal")
    np.testing.assert_allclose(raw_overview.x, raw_focus.x)
    np.testing.assert_allclose(raw_overview.y, raw_focus.y)
    assert tuple(focus.layout.xaxis.range) != tuple(overview.layout.xaxis.range or ())


def test_review_persistence_uses_synthetic_workbook(tmp_path):
    target = tmp_path / "synthetic_reference.xlsx"
    shutil.copy2(REF, target)
    result = run_pipeline(H5, target, 1001)
    queue = apply_review_state(build_review_queue(result, max_items=10), {})
    case = queue.iloc[0]
    persist_review_record(
        target, case,
        state="accepted",
        note="Synthetic demo review",
        action="Accept as reviewed",
        detail=f"{case['region_id']} accepted.",
    )
    records = load_review_records(target, 1001)
    assert records[str(case["case_key"])]["state"] == "accepted"
    assert load_persisted_audit(target, 1001)[0]["action"] == "Accept as reviewed"
    workbook = load_workbook(target, read_only=True)
    assert "reference_peaks" in workbook.sheetnames
    assert "ReviewActions" in workbook.sheetnames
    assert "ReviewAuditTrail" in workbook.sheetnames
    completed = apply_review_state(queue, records)
    counts = review_queue_counts(completed)
    assert counts["reviewed"] == 1
    exported = json.loads(report_json(result, completed, load_persisted_audit(target, 1001), records, 1001, str(target), str(H5)))
    assert exported["data_sources"]["synthetic_signal_h5"].endswith("synthetic_chromatograms.h5")


def test_optional_model_is_synthetic_only():
    artifact = joblib.load(MODEL)
    assert artifact["model_type"] == "random_forest_synthetic_demo"
    assert artifact["metrics"]["dataset"] == "public synthetic demo only"
    assert len(artifact["feature_columns"]) == 21
